import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils import get_column_letter
import pandas as pd

# Define the tickers for various stocks
tickers = {
    "ENEL": "ENEL.MI",
    "INTESA SANPAOLO": "ISP.MI",
    "POSTE ITALIANE": "PST.MI",
    "BANCO BPM": "BAMI.MI",
    "STELLANTIS": "STLAM.MI",
    "GENERALI": "G.MI"
}

# Create a dictionary to store prices
prices = {}

# Get the closing prices of the stocks
for company, ticker in tickers.items():
    stock = yf.Ticker(ticker)
    hist = stock.history(period="1d")
    if not hist.empty:
        price = hist['Close'].iloc[-1]
        prices[company] = price
    else:
        print(f"No data found for {company}")

# Path to the Excel file
file_path = r'.......'  # Replace with the correct path to your Excel file

# Try to open the existing Excel file
workbook = load_workbook(file_path)

# Check if the sheet with the stock data exists
sheet_name = 'Sheet1'  # Replace with the correct sheet name
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
else:
    sheet = workbook.active
    sheet.title = sheet_name

# Update the existing sheet with stock prices
for row in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=4).value
    if cell_value in prices:
        sheet.cell(row=row, column=6, value=prices[cell_value])  # Update 'Valore per azione'

# Create a new sheet for the charts
chart_sheet = workbook.create_sheet(title="Investment Charts")

# Generate and add charts to the new sheet
for idx, (company, ticker) in enumerate(tickers.items(), start=1):
    stock = yf.Ticker(ticker)
    hist = stock.history(period="6mo")  # Get data for the last 6 months
    if not hist.empty:
        # Remove timezone info from datetime
        hist.index = hist.index.tz_localize(None)
        
        # Add the data to the new sheet
        data_col = 2 * idx
        chart_sheet.cell(row=1, column=data_col-1, value=company)
        chart_sheet.cell(row=1, column=data_col, value="Close")
        for r, (date, close) in enumerate(zip(hist.index, hist['Close']), start=2):
            chart_sheet.cell(row=r, column=data_col-1, value=date)
            chart_sheet.cell(row=r, column=data_col, value=close)
        
        # Create a line chart for the stock price
        line_chart = LineChart()
        line_chart.title = f'{company} Stock Price'
        line_chart.y_axis.title = 'Price'
        line_chart.x_axis.title = 'Date'

        data = Reference(chart_sheet, min_col=data_col, min_row=2, max_row=r)
        dates = Reference(chart_sheet, min_col=data_col-1, min_row=2, max_row=r)
        line_chart.add_data(data, titles_from_data=False)
        line_chart.set_categories(dates)

        # Position the line chart in the sheet
        chart_sheet.add_chart(line_chart, f'{get_column_letter(data_col-1)}{r+2}')
    else:
        print(f"No historical data found for {company}")

# Check if the dividend sheet exists
dividend_sheet_name = 'Dividends'  # Replace with the correct sheet name
if dividend_sheet_name in workbook.sheetnames:
    dividend_sheet = workbook[dividend_sheet_name]
else:
    dividend_sheet = workbook.create_sheet(title=dividend_sheet_name)

# Generate a bar chart for dividends
dividend_chart = BarChart()
dividend_chart.title = "Dividends"
dividend_chart.y_axis.title = "Amount"
dividend_chart.x_axis.title = "Ticker"

dividend_data = Reference(dividend_sheet, min_col=4, min_row=2, max_row=dividend_sheet.max_row, max_col=5)
dividend_categories = Reference(dividend_sheet, min_col=2, min_row=2, max_row=dividend_sheet.max_row)
dividend_chart.add_data(dividend_data, titles_from_data=True)
dividend_chart.set_categories(dividend_categories)

# Position the bar chart in the sheet
chart_sheet.add_chart(dividend_chart, "A20")

# Save the Excel file
workbook.save(file_path)

print("Updated stock prices and generated charts in Excel:")
for company, price in prices.items():
    print(f"{company}: {price}")
